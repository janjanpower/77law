#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 讀取器 - Heroku 安全版本
修復 pandas 導入問題，確保在 Heroku 上正常運行
"""

import os
import sys
from pathlib import Path
from typing import List, Dict, Optional, Any, Tuple

# 安全導入 pandas
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
    print("✅ pandas 導入成功")
except ImportError:
    pd = None
    PANDAS_AVAILABLE = False
    print("⚠️ pandas 未安裝，Excel功能將受限")

# 安全導入其他 Excel 處理庫
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    openpyxl = None
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    xlrd = None
    XLRD_AVAILABLE = False


class ExcelReader:
    """Excel 讀取器 - 安全版本"""

    def __init__(self):
        """初始化 Excel 讀取器"""
        self.pandas_available = PANDAS_AVAILABLE
        self.openpyxl_available = OPENPYXL_AVAILABLE
        self.xlrd_available = XLRD_AVAILABLE

        if not any([PANDAS_AVAILABLE, OPENPYXL_AVAILABLE, XLRD_AVAILABLE]):
            print("⚠️ 警告: 沒有可用的 Excel 處理庫")

    def check_dependencies(self) -> Dict[str, bool]:
        """檢查依賴庫狀態"""
        return {
            'pandas': self.pandas_available,
            'openpyxl': self.openpyxl_available,
            'xlrd': self.xlrd_available
        }

    def read_excel(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        header_row: int = 0
    ) -> Optional[Any]:  # 改為 Any 避免 pd.DataFrame 類型錯誤
        """
        讀取 Excel 文件

        Args:
            file_path: Excel 文件路徑
            sheet_name: 工作表名稱（可選）
            header_row: 標題行索引

        Returns:
            DataFrame 或 None（如果讀取失敗）
        """
        try:
            if not os.path.exists(file_path):
                print(f"❌ 文件不存在: {file_path}")
                return None

            if not self.pandas_available:
                print("❌ pandas 不可用，無法讀取 Excel")
                return None

            # 使用 pandas 讀取 Excel
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header_row,
                engine='openpyxl' if self.openpyxl_available else None
            )

            print(f"✅ 成功讀取 Excel: {file_path}")
            print(f"   📊 形狀: {df.shape}")

            return df

        except Exception as e:
            print(f"❌ 讀取 Excel 失敗: {e}")
            return None

    def get_sheet_names(self, file_path: str) -> List[str]:
        """獲取 Excel 文件中的工作表名稱"""
        try:
            if not self.pandas_available:
                return []

            excel_file = pd.ExcelFile(file_path)
            return excel_file.sheet_names

        except Exception as e:
            print(f"❌ 獲取工作表名稱失敗: {e}")
            return []

    def read_excel_basic(self, file_path: str) -> Dict[str, Any]:
        """
        基本 Excel 讀取（不依賴 pandas）
        使用 openpyxl 作為後備方案
        """
        try:
            if not self.openpyxl_available:
                return {"error": "openpyxl 不可用"}

            from openpyxl import load_workbook

            wb = load_workbook(file_path)
            sheet_names = wb.sheetnames

            data = {}
            for sheet_name in sheet_names:
                ws = wb[sheet_name]

                # 讀取所有數據
                sheet_data = []
                for row in ws.iter_rows(values_only=True):
                    sheet_data.append(list(row))

                data[sheet_name] = sheet_data

            return {
                "success": True,
                "sheets": data,
                "sheet_names": sheet_names
            }

        except Exception as e:
            return {"error": str(e)}


# 為了向後兼容性，提供一個簡單的函數接口
def safe_read_excel(file_path: str) -> Optional[Any]:
    """安全的 Excel 讀取函數"""
    reader = ExcelReader()
    return reader.read_excel(file_path)


# 如果作為主程式運行，進行簡單測試
if __name__ == "__main__":
    reader = ExcelReader()
    print("📊 Excel 讀取器狀態:")
    deps = reader.check_dependencies()
    for lib, available in deps.items():
        status = "✅" if available else "❌"
        print(f"   {status} {lib}")
