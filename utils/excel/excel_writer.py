#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel寫入器 - 專責Excel檔案寫入功能
提供統一的Excel檔案寫入介面，支援多工作表和格式美化
"""

import os
from typing import List, Dict, Any, Optional
from pathlib import Path

# 安全的依賴導入
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from models.case_model import CaseData
from .exceptions import ExcelWriteError, ExcelDependencyError


class ExcelWriter:
    """Excel寫入器類別"""

    def __init__(self):
        """初始化Excel寫入器"""
        self._check_dependencies()

    def _check_dependencies(self) -> None:
        """檢查必要依賴"""
        if not PANDAS_AVAILABLE:
            raise ExcelDependencyError("pandas 不可用，無法寫入Excel檔案")
        if not OPENPYXL_AVAILABLE:
            raise ExcelDependencyError("openpyxl 不可用，無法寫入Excel檔案")

    def export_cases_to_excel(
        self,
        cases: List[CaseData],
        file_path: str,
        include_formatting: bool = True
    ) -> bool:
        """
        將案件資料匯出為Excel檔案

        Args:
            cases: 案件資料列表
            file_path: 輸出檔案路徑
            include_formatting: 是否包含格式美化

        Returns:
            匯出是否成功

        Raises:
            ExcelWriteError: 寫入失敗
        """
        try:
            if not cases:
                raise ExcelWriteError("沒有案件資料可匯出")

            # 準備資料
            data = self._prepare_case_data(cases)
            df = pd.DataFrame(data)

            # 建立目錄（如果不存在）
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            # 寫入Excel檔案
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='案件資料', index=False)

                if include_formatting:
                    self._apply_formatting(writer.sheets['案件資料'], df)

            print(f"✅ 成功匯出 {len(cases)} 筆案件到 {file_path}")
            return True

        except Exception as e:
            error_msg = f"匯出Excel失敗: {str(e)}"
            print(f"❌ {error_msg}")
            raise ExcelWriteError(error_msg)

    def export_cases_by_type(
        self,
        cases: List[CaseData],
        file_path: str,
        include_formatting: bool = True
    ) -> bool:
        """
        將案件資料按類型分工作表匯出

        Args:
            cases: 案件資料列表
            file_path: 輸出檔案路徑
            include_formatting: 是否包含格式美化

        Returns:
            匯出是否成功
        """
        try:
            if not cases:
                raise ExcelWriteError("沒有案件資料可匯出")

            # 按類型分組
            cases_by_type = self._group_cases_by_type(cases)

            # 建立目錄（如果不存在）
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for case_type, type_cases in cases_by_type.items():
                    # 準備該類型的資料
                    data = self._prepare_case_data(type_cases)
                    df = pd.DataFrame(data)

                    # 清理工作表名稱
                    sheet_name = self._sanitize_sheet_name(case_type)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                    if include_formatting:
                        self._apply_formatting(writer.sheets[sheet_name], df)

                    print(f"    ✅ {case_type}: {len(type_cases)} 筆案件")

            print(f"✅ 成功匯出 {len(cases)} 筆案件到 {file_path}（按類型分工作表）")
            return True

        except Exception as e:
            error_msg = f"分類匯出Excel失敗: {str(e)}"
            print(f"❌ {error_msg}")
            raise ExcelWriteError(error_msg)

    def create_case_info_excel(
        self,
        case_data: CaseData,
        file_path: str,
        include_progress: bool = True
    ) -> bool:
        """
        為單一案件建立詳細資訊Excel檔案

        Args:
            case_data: 案件資料
            file_path: 輸出檔案路徑
            include_progress: 是否包含進度追蹤工作表

        Returns:
            建立是否成功
        """
        try:
            # 建立目錄（如果不存在）
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 基本資訊工作表
                basic_info = self._prepare_basic_info_data(case_data)
                basic_df = pd.DataFrame(basic_info, columns=['項目', '內容'])
                basic_df.to_excel(writer, sheet_name='基本資訊', index=False)

                # 詳細資訊工作表
                detail_info = self._prepare_detail_info_data(case_data)
                detail_df = pd.DataFrame(detail_info, columns=['項目', '內容'])
                detail_df.to_excel(writer, sheet_name='詳細資訊', index=False)

                # 進度追蹤工作表
                if include_progress and hasattr(case_data, 'progress_stages') and case_data.progress_stages:
                    progress_info = self._prepare_progress_info_data(case_data)
                    progress_df = pd.DataFrame(progress_info, columns=['階段', '日期', '備註', '時間'])
                    progress_df.to_excel(writer, sheet_name='進度追蹤', index=False)

                    # 格式化進度追蹤工作表
                    self._apply_formatting(writer.sheets['進度追蹤'], progress_df)

                # 格式化基本和詳細資訊工作表
                self._apply_info_formatting(writer.sheets['基本資訊'], basic_df)
                self._apply_info_formatting(writer.sheets['詳細資訊'], detail_df)

            print(f"✅ 成功建立案件資訊Excel: {file_path}")
            return True

        except Exception as e:
            error_msg = f"建立案件資訊Excel失敗: {str(e)}"
            print(f"❌ {error_msg}")
            raise ExcelWriteError(error_msg)

    def append_cases_to_excel(
        self,
        cases: List[CaseData],
        file_path: str,
        sheet_name: str = '案件資料'
    ) -> bool:
        """
        將案件資料追加到現有Excel檔案

        Args:
            cases: 案件資料列表
            file_path: Excel檔案路徑
            sheet_name: 工作表名稱

        Returns:
            追加是否成功
        """
        try:
            if not cases:
                raise ExcelWriteError("沒有案件資料可追加")

            # 準備新資料
            new_data = self._prepare_case_data(cases)
            new_df = pd.DataFrame(new_data)

            if os.path.exists(file_path):
                # 讀取現有資料
                try:
                    existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                    # 合併資料
                    combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                except Exception:
                    # 如果讀取失敗，使用新資料
                    combined_df = new_df
            else:
                combined_df = new_df

            # 寫入合併後的資料
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
                self._apply_formatting(writer.sheets[sheet_name], combined_df)

            print(f"✅ 成功追加 {len(cases)} 筆案件到 {file_path}")
            return True

        except Exception as e:
            error_msg = f"追加案件到Excel失敗: {str(e)}"
            print(f"❌ {error_msg}")
            raise ExcelWriteError(error_msg)

    def _prepare_case_data(self, cases: List[CaseData]) -> List[Dict[str, Any]]:
        """準備案件資料用於匯出"""
        data = []
        for case in cases:
            data.append({
                '案件編號': getattr(case, 'case_id', '') or '',
                '案件類型': getattr(case, 'case_type', '') or '',
                '當事人': getattr(case, 'client', '') or '',
                '案由': getattr(case, 'case_reason', '') or '',
                '案號': getattr(case, 'case_number', '') or '',
                '負責法院': getattr(case, 'court', '') or '',
                '負責股別': getattr(case, 'division', '') or '',
                '對造': getattr(case, 'opposing_party', '') or '',
                '委任律師': getattr(case, 'lawyer', '') or '',
                '法務': getattr(case, 'legal_affairs', '') or '',
                '進度追蹤': getattr(case, 'progress', '') or '待處理'
            })
        return data

    def _prepare_basic_info_data(self, case_data: CaseData) -> List[List[str]]:
        """準備基本資訊資料"""
        return [
            ['案件編號', getattr(case_data, 'case_id', '') or ''],
            ['案件類型', getattr(case_data, 'case_type', '') or ''],
            ['當事人', getattr(case_data, 'client', '') or ''],
            ['委任律師', getattr(case_data, 'lawyer', '') or ''],
            ['法務', getattr(case_data, 'legal_affairs', '') or ''],
            ['進度追蹤', getattr(case_data, 'progress', '') or '待處理']
        ]

    def _prepare_detail_info_data(self, case_data: CaseData) -> List[List[str]]:
        """準備詳細資訊資料"""
        return [
            ['案由', getattr(case_data, 'case_reason', '') or ''],
            ['案號', getattr(case_data, 'case_number', '') or ''],
            ['負責法院', getattr(case_data, 'court', '') or ''],
            ['負責股別', getattr(case_data, 'division', '') or ''],
            ['對造', getattr(case_data, 'opposing_party', '') or ''],
            ['建立日期', getattr(case_data, 'created_at', '') or ''],
            ['更新日期', getattr(case_data, 'updated_at', '') or '']
        ]

    def _prepare_progress_info_data(self, case_data: CaseData) -> List[List[str]]:
        """準備進度追蹤資料"""
        progress_data = []
        if hasattr(case_data, 'progress_stages') and case_data.progress_stages:
            for stage_name, stage_info in case_data.progress_stages.items():
                if isinstance(stage_info, dict):
                    progress_data.append([
                        stage_name,
                        stage_info.get('date', ''),
                        stage_info.get('notes', ''),
                        stage_info.get('time', '')
                    ])
                else:
                    progress_data.append([stage_name, '', str(stage_info), ''])

        if not progress_data:
            progress_data.append(['待處理', '', '', ''])

        return progress_data

    def _group_cases_by_type(self, cases: List[CaseData]) -> Dict[str, List[CaseData]]:
        """按案件類型分組"""
        grouped = {}
        for case in cases:
            case_type = getattr(case, 'case_type', '未知')
            if case_type not in grouped:
                grouped[case_type] = []
            grouped[case_type].append(case)
        return grouped

    def _sanitize_sheet_name(self, name: str) -> str:
        """清理工作表名稱"""
        # Excel工作表名稱限制
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        clean_name = str(name)

        for char in invalid_chars:
            clean_name = clean_name.replace(char, '_')

        # 限制長度
        return clean_name[:30] if len(clean_name) > 30 else clean_name

    def _apply_formatting(self, worksheet, df: pd.DataFrame) -> None:
        """套用基本格式化"""
        try:
            # 標題列格式
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            # 格式化標題列
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 調整欄寬
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # 凍結標題列
            worksheet.freeze_panes = 'A2'

        except Exception as e:
            print(f"⚠️ 套用格式化時發生錯誤: {e}")

    def _apply_info_formatting(self, worksheet, df: pd.DataFrame) -> None:
        """套用資訊工作表格式化"""
        try:
            # 項目欄位格式
            item_font = Font(bold=True)
            item_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

            # 格式化項目欄位
            for row_num in range(1, len(df) + 2):  # +2 因為包含標題列
                cell = worksheet.cell(row=row_num, column=1)
                if row_num == 1:  # 標題列
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                else:  # 項目列
                    cell.font = item_font
                    cell.fill = item_fill

            # 調整欄寬
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 30

        except Exception as e:
            print(f"⚠️ 套用資訊格式化時發生錯誤: {e}")

    @staticmethod
    def get_dependency_status() -> Dict[str, bool]:
        """取得依賴狀態"""
        return {
            'pandas': PANDAS_AVAILABLE,
            'openpyxl': OPENPYXL_AVAILABLE
        }